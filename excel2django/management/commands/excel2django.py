from django.apps import apps
from django.contrib.contenttypes.fields import GenericRelation
from django.core.management.base import BaseCommand
from django.db import models, transaction
from django.forms.models import model_to_dict
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


class IgnoreObject(Exception):
    pass


class Command(BaseCommand):
    help = "Import data from excel into django objects"

    def add_arguments(self, parser):
        parser.add_argument("import_file")

        parser.add_argument(
            "--rows",
            action="append",
            help="""Specify range of rows with <START>:<END>. The index starts at 1 and the range includes the element at index END.
Examples:
--rows 10:20 # all rows from 10 to 20 (including 20)
--rows 2: # all rows except the header row
--rows 2:-1 # All rows except the first and the last one

rows can be specified multiple times, all specified ranges will be imported.
If no argument for rows is given all rows will be imported
""",
        )

        parser.add_argument(
            "--field",
            action="append",
            help="""Define how the value of a field should be extracted from a row. The base syntax is:
    <FIELD-SPECIFICATION>:<VALUE-EXPRESSION>

The syntax for FIELD-SPECIFICATION is:
    <APP_NAME>.<MODEL_NAME>.<FIELD_NAME>

In addition different instances of the same model can be distinguished by adding an arbitrary group-name:
    <APP_NAME>.<MODEL_NAME>.<GROUP_NAME>.<FIELD_NAME>

If the field should be used as component of the natural key the character + should be prefixed:
    +<APP_NAME>.<MODEL_NAME>.<FIELD_NAME>

The FIELD-SPECIFICATION must be unique over all --field arguments except for One-To-Many or Many-To-Many fields.

The value of VALUE-EXPRESSION will be evaluated as python expression.
All columns are available as variables with the according volumn name like A, B, AA, etc.
If the column value is a string, whitespace will be stripped.
Apart from the python builtins there are a few helper functions help with transformations:
    ref(<APP_NAME>.<MODEL_NAME>, <FIELD1_NAME>=<VALUE1>, <FIELD2_NAME>=<VALUE2>) # converts the given VALUEs into a reference to <APP_NAME>.<MODEL_NAME> passing all arguments to the get-method of the manager.
    vmap(<INPUT>, (<SRC>, <DST>), ...) # Returns first DST[i] where INPUT == SRC[i], returns <INPUT> if <INPUT> does not match any <SRC>

Examples:
    --field example.Publisher.name:B \\
    --field example.Author.1.first_name:C \\
    --field +example.Author.1.email:D \\
    --field example.Author.2.first_name:E \\
    --field +example.Author.2.email:F \\
    --field example.Book.title:A \\
    --field 'example.Book.publisher:ref("example.Publisher", name=B)'
""",
        )
        parser.add_argument(
            "--sheet",
            help="Specify which sheet to use. Can be a 1-base index or the name of sheet. The default is the first sheet in the excel file.",
        )
        parser.add_argument(
            "-y",
            "--yes",
            action="store_true",
            default=False,
            help="Do not ask for confirmation of the data to import",
        )

    def handle(self, *args, **options):
        rowranges = [(None, None)]
        sheet = try_int(options.get("sheet") or 1)
        if options["rows"]:
            rowranges = combine_ranges(
                (
                    (
                        try_int(r.split(":", 1)[0], None),
                        try_int(r.split(":", 1)[1], None),
                    )
                    for r in options["rows"]
                )
            )

        workbook = load_workbook(options["import_file"], data_only=True, read_only=True)

        if isinstance(sheet, int):
            sheet = workbook.sheetnames[sheet - 1]
        worksheet = workbook.get_sheet_by_name(sheet)

        modeldefinitions = {}
        if "field" in options:
            for fieldarg in options["field"]:
                fieldspec, valueexpr = fieldarg.split(":", 1)
                app_model, field = fieldspec.rsplit(".", 1)
                is_natural_key = app_model.startswith("*")
                app_model = app_model[1:] if is_natural_key else app_model
                app_label, modelname = app_model.split(".", 1)
                if app_model not in modeldefinitions:
                    modeldefinitions[app_model] = {
                        "model": apps.get_model(app_label, modelname.split(".", 1)[0]),
                        "fields": {},
                    }
                modeldefinitions[app_model]["fields"][field] = {
                    "modelfield": modeldefinitions[app_model]["model"]._meta.get_field(
                        field
                    ),
                    "expression": valueexpr,
                    "is_natural_key": is_natural_key,
                }
        # if no fields from the command line have been marked as natural key, use all fields as part of the natural key
        for model in modeldefinitions:
            if not any(
                [
                    f["is_natural_key"]
                    for f in modeldefinitions[model]["fields"].values()
                ]
            ):
                for f in modeldefinitions[model]["fields"].values():
                    f["is_natural_key"] = True

        errors = []
        with transaction.atomic():
            for r in rowranges:
                for row in worksheet.iter_rows(min_row=r[0], max_row=r[1]):
                    rowcontext = {
                        get_column_letter(i): c.value for i, c in enumerate(row, 1)
                    }
                    # strip whitespace from strings
                    for col in rowcontext:
                        if isinstance(rowcontext[col], str):
                            rowcontext[col] = rowcontext[col].strip()
                    objects = []
                    for model in model_import_order(modeldefinitions):
                        try:
                            newversion, oldversion = importinstance(
                                modeldefinitions[model], rowcontext
                            )
                        except IgnoreObject:
                            pass
                        if newversion is not None:
                            objects.append((newversion, oldversion))
                    print(f"Objects from row {getattr(row[0], 'row', '??')}:")
                    for newversion, oldversion in objects:
                        newdict = model_to_dict(newversion)
                        changes = newdict
                        # simple object diff
                        if oldversion:
                            olddict = model_to_dict(oldversion)
                            changedfields = [
                                k
                                for k in (olddict.keys() ^ newdict.keys())
                                if newdict[k] != olddict[k]
                            ]
                            changes = ""
                            if changedfields:
                                changes += str({k: newdict[k] for k in changedfields})
                            if olddict.keys() - newdict.keys():
                                changes += (
                                    f" removed: {olddict.keys() - newdict.keys()}"
                                )
                            if newdict.keys() - olddict.keys():
                                changes += f" added: {newdict.keys() - olddict.keys()}"
                        if changes:
                            print(
                                f"    {newversion} ({'new' if not oldversion else 'updated'}) {changes}"
                            )
                        else:
                            print(f"    No changes for {newversion}")
            if "yes" not in options:
                do = input("Do you want to save the import? [Y/n] ")
                if do.lower() not in ["", "y"]:
                    raise Exception("Import cancelled by user")

        if errors:
            print("Errors:")
            for error in errors:
                print(f"  {error}")


def importinstance(model, rowcontext):
    naturalkey_values = {
        f["modelfield"].name: _extract_fieldvalue(f, rowcontext)
        for f in model["fields"].values()
        if f["is_natural_key"]
    }
    if not all(naturalkey_values.values()):
        return None, False
    default_values = {
        f["modelfield"].name: _extract_fieldvalue(f, rowcontext)
        for f in model["fields"].values()
        if not f["is_natural_key"]
        and not isinstance(
            f["modelfield"],
            (models.fields.reverse_related.ForeignObjectRel, GenericRelation),
        )
    }
    oldobject = model["model"].objects.filter(**naturalkey_values).first()
    newobject, _ = model["model"].objects.update_or_create(
        defaults=default_values, **naturalkey_values
    )
    for f in model["fields"].values():
        if isinstance(
            f["modelfield"],
            (models.fields.reverse_related.ForeignObjectRel, GenericRelation),
        ):
            object_list = _extract_fieldvalue(f, rowcontext)
            if object_list:
                newobject.save()
                if isinstance(object_list[0], models.Model):
                    getattr(newobject, f["modelfield"].name).add(
                        *object_list, bulk=False
                    )
                else:
                    for args in object_list:
                        # poor duplication detection...
                        obj = (
                            getattr(newobject, f["modelfield"].name)
                            .filter(**args)
                            .first()
                        )
                        if not obj:
                            getattr(newobject, f["modelfield"].name).create(**args)

    return newobject, oldobject


def _extract_fieldvalue(field, rowcontext):
    return eval(field["expression"], {"ref": ref, "vmap": vmap}, rowcontext)


def combine_ranges(ranges):
    combined = []
    for r in sorted(ranges):
        if not combined:
            combined.append(r)
        elif not range_overlap(combined[-1][0], combined[-1][1], r[0], r[1]):
            combined.append(r)
        elif r[1] > combined[-1][1]:
            combined[-1][1] = r[1]
    return combined


def range_overlap(start1, end1, start2, end2):
    def contains(start, end, i):
        if start is None and end is None:
            return True
        if start is None:
            return i <= end
        if end is None:
            return start <= i
        return start <= i <= end

    return (
        contains(start1, end1, start2)
        or contains(start1, end1, end2)
        or contains(start2, end2, start1)
        or contains(start2, end2, end1)
    )


def model_import_order(modeldefinitions):
    # TODO: automatically detect correct order to satisfy all uses of "ref" (referenced objects in the same line should be created before their referers)
    return modeldefinitions.keys()


def try_int(v, default=None):
    try:
        return int(v)
    except (TypeError, ValueError):
        return v if default is None else default


# transformation functions ------------------------------------------------------


def ref(modelname, **kwargs):
    return apps.get_model(*modelname.split(".")).objects.filter(**kwargs).first()


def noempty(value):
    if value == "":
        raise IgnoreObject()
    return value


def vmap(_input, *mappings):
    for key, value in mappings:
        if _input == key:
            return value
    return _input
